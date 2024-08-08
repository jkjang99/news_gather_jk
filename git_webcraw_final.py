import requests
from bs4 import BeautifulSoup
from datetime import datetime
import time
from openpyxl import Workbook
import json
import pandas as pd
# import telegram
from telegram import Bot
import asyncio
import re
from difflib import SequenceMatcher
from openai import OpenAI
import os
import streamlit as st

# 설정 파일 읽기
script_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(script_dir, 'config.json')

with open(config_path, 'r', encoding='utf-8') as config_file:
    config = json.load(config_file)

working_dir = config.get('working_dir', script_dir)
client = OpenAI(api_key=config['openai_api_key'])

def clean_text(text):
    text = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)
    text = re.sub(r'[^\w\s\.,;:!?]', '', text)
    return text

def search_naver_news(keyword):
    base_url = "https://search.naver.com/search.naver?where=news&query="
    sort_option = "&sort=1"
    period_option = "&pd=4"
    start_option = "&start="

    results = []
    start = 1
    while True:
        search_url = base_url + requests.utils.quote(keyword) + sort_option + period_option + start_option + str(start)

        response = requests.get(search_url, headers={'User-Agent': 'Mozilla/5.0'})
        soup = BeautifulSoup(response.text, 'html.parser')
        
        news_items = soup.select('li.bx')
        
        if not news_items:
            break
        
        for item in news_items:
            title_elem = item.select_one('a.news_tit')
            press_elem = item.select_one('a.info.press')
            
            info_elements = item.select('.info_group span.info')
            
            news_time = "시간 정보 없음"
            for info in info_elements:
                if not info.find('i') and any(char.isdigit() for char in info.text):
                    news_time = info.text.strip()
                    break

            if title_elem and press_elem:
                news_title = title_elem.text.strip()
                news_url = title_elem['href']
                news_press = press_elem.text.strip()
                
                results.append((news_title, news_url, news_press, news_time))
        
        if len(news_items) < 10:
            break
        
        start += 10
        time.sleep(1)
    
    return results

def get_article_content(url):
    try:
        response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
        soup = BeautifulSoup(response.text, 'html.parser')
        
        possible_content_classes = ['#articleBodyContents', '#articeBody', '.article_body', '#newsEndContents', '.article-body', 'article-body','article-Body','news_article','articlecontent' ,'article_body', '#article-body']
        
        content = "기사 내용을 찾을 수 없습니다."
        for class_or_id in possible_content_classes:
            article_body = soup.select_one(class_or_id)
            if article_body:
                content = article_body.get_text(strip=True)
                break
        
        if content == "기사 내용을 찾을 수 없습니다.":
            paragraphs = soup.find_all('p')
            if paragraphs:
                content = ' '.join([p.get_text(strip=True) for p in paragraphs])
        
        return clean_text(content)
    except Exception as e:
        print(f"기사 내용 가져오기 실패: {url}")
        print(f"에러: {str(e)}")
        return "기사 내용을 가져오는 중 오류가 발생했습니다."

def analyze_article_importance(content):
    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 중대뉴스 판별사 입니다. 아래 기사가 진짜 사망사고에 대한 발생내역에 대한 뉴스를 담고 있는지를 판별해주세요. 본문기사의 내용에 '사망', '숨진', '사고사'라는 단어가 없는 경우는 중요하지 않은 기사로 판단할 수 있어. 다음 정보도 추출해주세요: 중요여부(true/false), 사고일시, 지역, 업체명, 사고내역. 각 항목을 새로운 줄에 'item: value' 형식으로 작성해주세요."},
                {"role": "user", "content": content}
            ],
            temperature=0,
            max_tokens=150
        )
        result = response.choices[0].message.content
        
        lines = result.split('\n')
        is_important = 'true' in lines[0].lower()
        
        accident_date = ''
        location = ''
        company = ''
        accident_details = ''
        
        for line in lines[1:]:
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip().lower()
                value = value.strip()
                if '사고일시' in key:
                    accident_date = value
                elif '지역' in key:
                    location = value
                elif '업체명' in key:
                    company = value
                elif '사고내역' in key:
                    accident_details = value
        
        return is_important, accident_date, location, company, accident_details
    except Exception as e:
        print(f"GPT 분석 중 오류 발생: {str(e)}")
        return False, '', '', '', ''

def calculate_similarity(text1, text2):
    return SequenceMatcher(None, text1, text2).ratio()

def run_news_analysis(keywords):
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"naver_news_results_{current_time}.xlsx"
    excel_filepath = os.path.join(working_dir, excel_filename)

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "전체 뉴스 결과"
    ws_unique = wb.create_sheet(title="중복 제거된 뉴스 결과")
    ws_removed = wb.create_sheet(title="제거된 뉴스 결과")

    headers = ['순번', '뉴스기사명', 'URL', '매체명', '기사내용', '키워드', '시간', '중요도', '사고일시', '지역', '업체명', '사고내역', '제거 이유']
    ws_all.append(headers)
    ws_unique.append(headers)
    ws_removed.append(headers)

    total_count = 0
    unique_count = 0
    removed_count = 0
    processed_urls = set()
    unique_articles = []

    for keyword in keywords:
        news_results = search_naver_news(keyword)
        
        for index, (title, url, press, newstime) in enumerate(news_results, start=1):
            if url in processed_urls:
                continue
            
            processed_urls.add(url)
            total_count += 1
            content = get_article_content(url)
            
            title = clean_text(title)
            press = clean_text(press)
            newstime = clean_text(newstime)
            keyword = clean_text(keyword)
            
            ws_all.append([total_count, title, url, press, content, keyword, newstime, '', ''])
            
            if len(content) <= 200:
                removed_count += 1
                ws_removed.append([total_count, title, url, press, content, keyword, newstime, '', '짧은 기사'])
                continue
            
            is_duplicate = False
            for unique_article in unique_articles:
                title_similarity = calculate_similarity(title, unique_article[1])
                content_similarity = calculate_similarity(content, unique_article[4])
                url_similarity = calculate_similarity(content, unique_article[2])
                region_similarity = calculate_similarity(content, unique_article[9])
                if title_similarity > 0.4 or content_similarity > 0.4 or region_similarity > 0.5 or url_similarity > 0.9:
                    is_duplicate = True
                    removed_count += 1
                    ws_removed.append([total_count, title, url, press, content, keyword, newstime, '', '유사한 기사'])
                    break
            
            if not is_duplicate:
                unique_count += 1
                is_important, accident_date, location, company, accident_details = analyze_article_importance(content)
                if not is_important:
                    removed_count += 1
                    ws_removed.append([total_count, title, url, press, content, keyword, newstime, 'False', accident_date, location, company, accident_details, '중요하지 않은 기사'])
                    continue
                unique_articles.append([unique_count, title, url, press, content, keyword, newstime, 'True', accident_date, location, company, accident_details, ''])
                ws_unique.append([unique_count, title, url, press, content, keyword, newstime, 'True', accident_date, location, company, accident_details, ''])
            
            time.sleep(1)

    wb.save(excel_filepath)
    
    return excel_filepath, total_count, unique_count, removed_count

# async def send_telegram_messages(excel_filepath):
#     bot = telegram.Bot(token=config['telegram_bot_token'])
#     channel_id = config['telegram_channel_id']

#     df = pd.read_excel(excel_filepath, sheet_name="중복 제거된 뉴스 결과")

#     for index, row in df.iterrows():
#         message = f"기사발생시간: {row['시간']}\n"
#         message += f"뉴스기사명: {row['뉴스기사명']}\n"
#         message += f"URL: {row['URL']}\n"
#         message += f"사고요약: {row['사고내역']}\n\n"
        
#         try:
#             await bot.send_message(chat_id=channel_id, text=message, parse_mode='HTML')
#             st.write(f"메시지 전송 완료: {row['뉴스기사명']}")
#         except telegram.error.TelegramError as e:
#             st.write(f"메시지 전송 실패: {e}")

#     st.write("모든 메시지가 전송되었습니다.")



# def main():
#     st.title("뉴스 분석 및 텔레그램 메시지 전송")

#     keywords = st.text_input("키워드를 입력하세요 (쉼표로 구분)", "중대재해사망사고")
#     keywords = [keyword.strip() for keyword in keywords.split(',')]

#     if st.button("뉴스 분석 시작"):
#         with st.spinner("뉴스를 분석 중입니다..."):
#             excel_filepath, total_count, unique_count, removed_count = run_news_analysis(keywords)

#         st.success("뉴스 분석이 완료되었습니다!")
#         st.write(f"Excel 파일이 생성되었습니다: {excel_filepath}")
#         st.write(f"총 기사 수: {total_count}")
#         st.write(f"중복 제거 후 기사 수: {unique_count}")
#         st.write(f"제거된 기사 수: {removed_count}")

#         if st.button("텔레그램으로 메시지 전송"):
#             with st.spinner("텔레그램으로 메시지를 전송 중입니다..."):
#                 asyncio.run(send_telegram_messages(excel_filepath))
#             st.success("텔레그램 메시지 전송이 완료되었습니다!")

# if __name__ == "__main__":
#     main()

# sync로 변경수정(8/7 1시) 
#def send_telegram_messages_sync(excel_filepath):
#     bot = Bot(token=config['telegram_bot_token'])
#     channel_id = config['telegram_channel_id']

#     df = pd.read_excel(excel_filepath, sheet_name="중복 제거된 뉴스 결과")

#     for index, row in df.iterrows():
#         message = f"기사발생시간: {row['시간']}\n"
#         message += f"뉴스기사명: {row['뉴스기사명']}\n"
#         message += f"URL: {row['URL']}\n"
#         message += f"사고요약: {row['사고내역']}\n\n"
        
#         try:
#             bot.send_message(chat_id=channel_id, text=message, parse_mode='HTML')
#             st.write(f"메시지 전송 완료: {row['뉴스기사명']}")
#         except telegram.error.TelegramError as e:
#             st.write(f"메시지 전송 실패: {e}")

#     st.write("모든 메시지가 전송되었습니다.")

# def main():
#     st.title("뉴스 분석 및 텔레그램 메시지 전송")

#     if 'excel_filepath' not in st.session_state:
#         st.session_state.excel_filepath = None

#     keywords = st.text_input("키워드를 입력하세요 (쉼표로 구분)", "중대재해사망사고")
#     keywords = [keyword.strip() for keyword in keywords.split(',')]

#     if st.button("뉴스 분석 시작"):
#         with st.spinner("뉴스를 분석 중입니다..."):
#             excel_filepath, total_count, unique_count, removed_count = run_news_analysis(keywords)
#             st.session_state.excel_filepath = excel_filepath

#         st.success("뉴스 분석이 완료되었습니다!")
#         st.write(f"Excel 파일이 생성되었습니다: {excel_filepath}")
#         st.write(f"총 기사 수: {total_count}")
#         st.write(f"중복 제거 후 기사 수: {unique_count}")
#         st.write(f"제거된 기사 수: {removed_count}")

#     if st.session_state.excel_filepath and st.button("텔레그램으로 메시지 전송"):
#         with st.spinner("텔레그램으로 메시지를 전송 중입니다..."):
#             send_telegram_messages_sync(st.session_state.excel_filepath)
#         st.success("텔레그램 메시지 전송이 완료되었습니다!")

# if __name__ == "__main__":
#     main()

# 다시 async 방식으로 수정(8/7 2시) 
async def send_telegram_message(bot, channel_id, message):
    await bot.send_message(chat_id=channel_id, text=message, parse_mode='HTML')

def send_telegram_messages_sync(excel_filepath):
    bot = Bot(token=config['telegram_bot_token'])
    channel_id = config['telegram_channel_id']

    df = pd.read_excel(excel_filepath, sheet_name="중복 제거된 뉴스 결과")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    for index, row in df.iterrows():
        message = f"기사발생시간: {row['시간']}\n"
        message += f"뉴스기사명: {row['뉴스기사명']}\n"
        message += f"URL: {row['URL']}\n"
        message += f"사고요약: {row['사고내역']}\n\n"
        
        try:
            loop.run_until_complete(send_telegram_message(bot, channel_id, message))
            st.write(f"메시지 전송 완료: {row['뉴스기사명']}")
        except Exception as e:
            st.write(f"메시지 전송 실패: {e}")

    loop.close()
    st.write("모든 메시지가 전송되었습니다.")

def main():
    st.title("뉴스 분석 및 텔레그램 메시지 전송")

    if 'excel_filepath' not in st.session_state:
        st.session_state.excel_filepath = None

    keywords = st.text_input("키워드를 입력하세요 (쉼표로 구분)", "중대재해사망사고")
    keywords = [keyword.strip() for keyword in keywords.split(',')]

    if st.button("뉴스 분석 시작"):
        with st.spinner("뉴스를 분석 중입니다..."):
            excel_filepath, total_count, unique_count, removed_count = run_news_analysis(keywords)
            st.session_state.excel_filepath = excel_filepath

        st.success("뉴스 분석이 완료되었습니다!")
        st.write(f"Excel 파일이 생성되었습니다: {excel_filepath}")
        st.write(f"총 기사 수: {total_count}")
        st.write(f"중복 제거 후 기사 수: {unique_count}")
        st.write(f"제거된 기사 수: {removed_count}")

    if st.session_state.excel_filepath and st.button("텔레그램으로 메시지 전송"):
        with st.spinner("텔레그램으로 메시지를 전송 중입니다..."):
            send_telegram_messages_sync(st.session_state.excel_filepath)
        st.success("텔레그램 메시지 전송이 완료되었습니다!")

if __name__ == "__main__":
    main()