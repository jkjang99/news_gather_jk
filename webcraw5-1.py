import requests
from bs4 import BeautifulSoup
from datetime import datetime
import time
from openpyxl import Workbook
import json
import pandas as pd # 8/1 밤 추가
import telegram # 8/1 밤 추가
import asyncio # 8/1 밤 추가
import re
from difflib import SequenceMatcher
from openai import OpenAI
import os

# # working_dir이 존재하지 않으면 생성합니다
# if not os.path.exists(working_dir):
#     os.makedirs(working_dir)

# # 설정 파일 읽기
# with open('C:/Users/SKTelecom/Documents/AI_Frontier3기/Event_logging/config.json', 'r',encoding='utf-8') as config_file: # 8/7 수정 폴더위치오류수정
#     config = json.load(config_file)

# 스크립트 파일의 디렉토리 경로를 가져옵니다
script_dir = os.path.dirname(os.path.abspath(__file__))

# config.json 파일의 경로를 구성합니다
config_path = os.path.join(script_dir, 'config.json')

# config 파일을 읽습니다
with open(config_path, 'r', encoding='utf-8') as config_file:
    config = json.load(config_file)

# working_dir를 가져옵니다
working_dir = config.get('working_dir', script_dir)

client = OpenAI(api_key=config['openai_api_key'])

def clean_text(text):
    # 컨트롤 문자 제거
    text = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)
    # 특수 문자 제거 (필요에 따라 조정)
    text = re.sub(r'[^\w\s\.,;:!?]', '', text)
    return text

def search_naver_news(keyword):
    base_url = "https://search.naver.com/search.naver?where=news&query="
    sort_option = "&sort=1"  # 1: 최신순, 0: 관련도순
    period_option = "&pd=4"  # 4: 1일, pd=2: 1주일, pd=1: 1개월
    start_option = "&start="

    results = []
    start = 1
    while True:
        search_url = base_url + requests.utils.quote(keyword) + sort_option + period_option + start_option + str(start)

        response = requests.get(search_url, headers={'User-Agent': 'Mozilla/5.0'})
        soup = BeautifulSoup(response.text, 'html.parser')
        
        news_items = soup.select('li.bx')
        
        if not news_items:
            break  # 더 이상 결과가 없으면 반복 중단
        
        for item in news_items:
            title_elem = item.select_one('a.news_tit')
            press_elem = item.select_one('a.info.press')
            
            info_elements = item.select('.info_group span.info')
            
            news_time = "시간 정보 없음"
            for info in info_elements:
                if not info.find('i') and any(char.isdigit() for char in info.text):
                    news_time = info.text.strip()
                    break

            if title_elem and press_elem:
                news_title = title_elem.text.strip()
                news_url = title_elem['href']
                news_press = press_elem.text.strip()
                
                results.append((news_title, news_url, news_press, news_time))
        
        if len(news_items) < 10:
            break  # 마지막 페이지라면 반복 중단
        
        start += 10  # 다음 페이지로 이동
        time.sleep(1)  # 너무 빠른 요청을 방지하기 위한 대기 시간
    
    return results

def get_article_content(url):
    try:
        response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
        soup = BeautifulSoup(response.text, 'html.parser')
        
        possible_content_classes = ['#articleBodyContents', '#articeBody', '.article_body', '#newsEndContents', '.article-body', 'article-body','article-Body','news_article','articlecontent' ,'article_body', '#article-body']
        
        content = "기사 내용을 찾을 수 없습니다."
        for class_or_id in possible_content_classes:
            article_body = soup.select_one(class_or_id)
            if article_body:
                content = article_body.get_text(strip=True)
                break
        
        if content == "기사 내용을 찾을 수 없습니다.":
            paragraphs = soup.find_all('p')
            if paragraphs:
                content = ' '.join([p.get_text(strip=True) for p in paragraphs])
        
        return clean_text(content)
    except Exception as e:
        print(f"기사 내용 가져오기 실패: {url}")
        print(f"에러: {str(e)}")
        return "기사 내용을 가져오는 중 오류가 발생했습니다."

# def analyze_article_importance(content): : 8/1 오후 수정
#     try:
#         response = client.chat.completions.create(
#             model="gpt-4o-mini",
#             messages=[
#                 {"role": "system", "content": "당신은 중대뉴스 판별사 입니다. 아래 기사가 진짜 사망사고에 대한 발생내역에 대한 뉴스를 담고 있는지를 판별해주세요. 본문기사의 내용에 '사망', '숨진', '사고사'라는 단어가 없는 경우는 중요하지 않은 기사로 판단할 수 있어. 답변은 '중요여부: (true/false)'로 시작하고, 그 다음 줄에 '이유: (판단 이유)'를 작성해주세요."},
#                 {"role": "user", "content": content}
#             ],
#             temperature=0,
#             max_tokens=100
#         )
#         result = response.choices[0].message.content
        
#         # 텍스트 파싱
#         lines = result.split('\n')
#         is_important = 'true' in lines[0].lower()
#         reason = lines[1].split('이유:', 1)[1].strip() if len(lines) > 1 else '이유가 제공되지 않았습니다.'
        
#         return is_important, reason
#     except Exception as e:
#         print(f"GPT 분석 중 오류 발생: {str(e)}")
#         return False, f"분석 중 오류 발생: {str(e)}"
    
def analyze_article_importance(content):
    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                #     {"role": "system", "content": "당신은 중대뉴스 판별사 입니다. 아래 기사가 진짜 사망사고에 대한 발생내역에 대한 뉴스를 담고 있는지를 판별해주세요. 본문기사의 내용에 '사망', '숨진', '사고사'라는 단어가 없는 경우는 중요하지 않은 기사로 판단할 수 있어. 답변은 '중요여부: (true/false)'로 시작하고, 그 다음 줄에 '이유: (판단 이유)'를 작성해주세요."},
                # {"role": "user", "content": content}
            #     {"role": "system", "content": "당신은 중대뉴스 판별사 입니다. 아래 기사가 진짜 사망사고에 대한 발생내역에 대한 뉴스를 담고 있는지를 판별해주세요. 본문기사의 내용에 '사망', '숨진', '사고사'라는 단어가 없는 경우는 중요하지 않은 기사로 판단할 수 있어. 다음 정보도 추출해주세요:중요여부(true/false), 사고일시, 지역, 업체명, 전체기사요약(100자 이내로 지역과 업체명이 있으면 포함해줘). 각 항목을 새로운 줄에 'item: value' 형식으로 작성해주세요."},
            #     {"role": "user", "content": content}
            # ],
                {"role": "system", "content": "당신은 중대뉴스 판별사 입니다. 아래 기사가 진짜 사망사고에 대한 발생내역에 대한 뉴스를 담고 있는지를 판별해주세요. 본문기사의 내용에 '사망', '숨진', '사고사'라는 단어가 없는 경우는 중요하지 않은 기사로 판단할 수 있어. 다음 정보도 추출해주세요: 중요여부(true/false), 사고일시, 지역, 업체명, 사고내역. 각 항목을 새로운 줄에 'item: value' 형식으로 작성해주세요."},
                {"role": "user", "content": content}
            ],
            temperature=0,
            max_tokens=150
        )
        result = response.choices[0].message.content
        
        # 결과 파싱
        lines = result.split('\n')
        is_important = 'true' in lines[0].lower()
        
        accident_date = ''
        location = ''
        company = ''
        accident_details = ''
        
        for line in lines[1:]:
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip().lower()
                value = value.strip()
                if '사고일시' in key:
                    accident_date = value
                elif '지역' in key:
                    location = value
                elif '업체명' in key:
                    company = value
                elif '사고내역' in key:
                    accident_details = value
        
        return is_important, accident_date, location, company, accident_details
    except Exception as e:
        print(f"GPT 분석 중 오류 발생: {str(e)}")
        return False, '', '', '', ''


def calculate_similarity(text1, text2):
    return SequenceMatcher(None, text1, text2).ratio()

# 키워드로 검색
# keywords = ["공사현장사망"]
keywords = ["중대재해사망사고"]

# 현재 날짜와 시간을 파일명에 포함
current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_filename = f"naver_news_results_{current_time}.xlsx"

# 전체 파일 경로를 생성합니다
excel_filepath = os.path.join(working_dir, excel_filename)

# Excel 워크북 생성
wb = Workbook()
ws_all = wb.active
ws_all.title = "전체 뉴스 결과"
ws_unique = wb.create_sheet(title="중복 제거된 뉴스 결과")
ws_removed = wb.create_sheet(title="제거된 뉴스 결과")

# # 헤더 작성
# headers = ['순번', '뉴스기사명', 'URL', '매체명', '기사내용', '키워드', '시간', 'GPT 분석 결과', '제거 이유']
# ws_all.append(headers)
# ws_unique.append(headers)
# ws_removed.append(headers)

# Excel 워크북 생성 부분 수정: 8/1 오후 수정
headers = ['순번', '뉴스기사명', 'URL', '매체명', '기사내용', '키워드', '시간', '중요도', '사고일시', '지역', '업체명', '사고내역', '제거 이유']
ws_all.append(headers)
ws_unique.append(headers)
ws_removed.append(headers)

total_count = 0
unique_count = 0
removed_count = 0
processed_urls = set()
unique_articles = []

for keyword in keywords:
    # print(f"\n키워드 '{keyword}'에 대한 검색 결과:")
    news_results = search_naver_news(keyword)
    
    for index, (title, url, press, newstime) in enumerate(news_results, start=1):
        if url in processed_urls:
            # print(f"중복 URL 건너뛰기: {title}")
            continue
        
        processed_urls.add(url)
        total_count += 1
        print(f"기사 '{title}' 처리 중...")
        content = get_article_content(url)
        
        # 모든 문자열 값을 정제
        title = clean_text(title)
        press = clean_text(press)
        newstime = clean_text(newstime)
        keyword = clean_text(keyword)
        
        # 전체 결과 시트에 추가
        ws_all.append([total_count, title, url, press, content, keyword, newstime, '', ''])
        
        # 기사 내용이 200자 이내인 경우 제거
        if len(content) <= 200:
            removed_count += 1
            ws_removed.append([total_count, title, url, press, content, keyword, newstime, '', '짧은 기사'])
            # print(f"짧은 기사 제거: {title}")
            continue
        
        # 유사도 검사
        is_duplicate = False
        for unique_article in unique_articles:
            title_similarity = calculate_similarity(title, unique_article[1]) or 0
            content_similarity = calculate_similarity(content, unique_article[4]) or 0
            url_similarity = calculate_similarity(content, unique_article[2]) or 0 # 8/1 오후: url의 유사도 추가
            region_similarity = calculate_similarity(content, unique_article[9]) or 0 # 8/1 오후: 지역의 유사도 추가
            # print(f"{unique_article[9]}")
            if title_similarity > 0.4 or content_similarity > 0.4 or region_similarity > 0.5 or url_similarity > 0.9:
                is_duplicate = True
                removed_count += 1
                ws_removed.append([total_count, title, url, press, content, keyword, newstime, '', '유사한 기사'])
                # print(f"유사한 기사 발견: {title}")
                break
        
        # if not is_duplicate: 8/1 최종 수정
        #     unique_count += 1
        #     is_important, analysis_result = analyze_article_importance(content)
        #     if not is_important:
        #         removed_count += 1
        #         ws_removed.append([total_count, title, url, press, content, keyword, newstime, analysis_result, '중요하지 않은 기사'])
        #         print(f"중요하지 않은 기사 제거: {title}")
        #         continue
        #     unique_articles.append([unique_count, title, url, press, content, keyword, newstime, analysis_result, ''])
        #     ws_unique.append([unique_count, title, url, press, content, keyword, newstime, analysis_result, ''])
        #     print(f"고유한 기사 추가: {title}")
        
# 결과 저장 부분 수정
        if not is_duplicate:
            unique_count += 1
            is_important, accident_date, location, company, accident_details = analyze_article_importance(content)
            if not is_important:
                removed_count += 1
                ws_removed.append([total_count, title, url, press, content, keyword, newstime, 'False', accident_date, location, company, accident_details, '중요하지 않은 기사'])
            #    print(f"중요하지 않은 기사 제거: {title}")
                continue
            unique_articles.append([unique_count, title, url, press, content, keyword, newstime, 'True', accident_date, location, company, accident_details, ''])
            ws_unique.append([unique_count, title, url, press, content, keyword, newstime, 'True', accident_date, location, company, accident_details, ''])
            # print(f"고유한 기사 추가: {title}")

        # print(f"순번: {total_count}")
        #print(f"뉴스기사명: {title}")
        # print(f"URL: {url}")
        # print(f"매체명: {press}")
        # print(f"검색어: {keyword}")
        # print(f"작성시간: {newstime}")
        #print("-" * 50)
        
        time.sleep(1)  # 웹사이트와 API에 부하를 주지 않기 위해 1초 대기

# Excel 파일 저장
wb.save(excel_filename)

print(f"\nExcel 파일이 생성되었습니다: {excel_filename}")
print(f"총 기사 수: {total_count}")
print(f"중복 제거 후 기사 수: {unique_count}")
print(f"제거된 기사 수: {removed_count}")

# Telegram 메시지 전송 함수
async def send_telegram_messages():
    # Telegram 봇 설정
    bot = telegram.Bot(token=config['telegram_bot_token'])
    channel_id = config['telegram_channel_id']

    # 엑셀 파일에서 "중복 제거된 뉴스 결과" 시트 읽기
    df = pd.read_excel(excel_filename, sheet_name="중복 제거된 뉴스 결과")

    for index, row in df.iterrows():
        message = f"기사발생시간: {row['시간']}\n"
        message += f"뉴스기사명: {row['뉴스기사명']}\n"
        message += f"URL: {row['URL']}\n"
        message += f"사고요약: {row['사고내역']}\n\n"
        
        try:
            # Telegram으로 메시지 전송
            await bot.send_message(chat_id=channel_id, text=message, parse_mode='HTML')
            print(f"메시지 전송 완료: {row['뉴스기사명']}")
        except telegram.error.TelegramError as e:
            print(f"메시지 전송 실패: {e}")

    print("모든 메시지가 전송되었습니다.")

# 메인 실행 부분
if __name__ == "__main__":
    # 기존 코드 실행 (뉴스 검색 및 Excel 파일 생성)
    # ...

    # Telegram 메시지 전송
    asyncio.run(send_telegram_messages())